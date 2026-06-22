import time
import urllib.request
import urllib.error
import concurrent.futures
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

URL = "https://concept60-1.onrender.com/"
CONCURRENT_USERS = 100
DURATION_SECONDS = 60

# We will store latencies (in milliseconds)
latencies = []

def fetch_url():
    start = time.time()
    try:
        req = urllib.request.Request(URL, headers={'User-Agent': 'LoadTester/1.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            response.read()
    except urllib.error.HTTPError as e:
        # We consider HTTP errors (like 429 Too Many Requests) as part of the test
        pass
    except Exception:
        pass
    end = time.time()
    return (end - start) * 1000  # ms

def worker(end_time):
    local_latencies = []
    while time.time() < end_time:
        lat = fetch_url()
        local_latencies.append(lat)
    return local_latencies

def run_load_test():
    print(f"Starting Load Test on {URL}")
    print(f"Users: {CONCURRENT_USERS}")
    print(f"Duration: {DURATION_SECONDS} seconds")
    
    start_time = time.time()
    end_time = start_time + DURATION_SECONDS
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=CONCURRENT_USERS) as executor:
        futures = [executor.submit(worker, end_time) for _ in range(CONCURRENT_USERS)]
        
        for future in concurrent.futures.as_completed(futures):
            latencies.extend(future.result())
            
    actual_duration = time.time() - start_time
    total_requests = len(latencies)
    
    if total_requests > 0:
        avg_time = sum(latencies) / total_requests
        min_time = min(latencies)
        max_time = max(latencies)
        rps = total_requests / actual_duration
    else:
        avg_time = min_time = max_time = rps = 0
        
    print(f"Total Requests: {total_requests}")
    print(f"RPS: {rps:.2f}")
    print(f"Min: {min_time:.2f}ms")
    print(f"Max: {max_time:.2f}ms")
    print(f"Avg: {avg_time:.2f}ms")
    
    # Save to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Load Test Results"
    
    # Styling
    def make_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def make_font(bold=False, color="000000", size=11): return Font(bold=bold, color=color, size=size, name="Calibri")
    def center_align(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
        
    headers = ["Metric", "Value"]
    for col, val in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=val)
        c.fill = make_fill("1A252F")
        c.font = make_font(bold=True, color="FFFFFF")
        c.alignment = center_align()
        c.border = thin_border()
        
    results = [
        ("Target URL", URL),
        ("Virtual Users", CONCURRENT_USERS),
        ("Duration (seconds)", round(actual_duration, 2)),
        ("Total Requests", total_requests),
        ("Requests Per Second (RPS)", round(rps, 2)),
        ("Average Response Time (ms)", round(avg_time, 2)),
        ("Min Response Time (ms)", round(min_time, 2)),
        ("Max Response Time (ms)", round(max_time, 2))
    ]
    
    for i, (metric, value) in enumerate(results, 2):
        c1 = ws.cell(row=i, column=1, value=metric)
        c1.border = thin_border()
        c1.font = make_font(bold=True)
        c1.fill = make_fill("EAF2F8" if i % 2 == 0 else "FFFFFF")
        
        c2 = ws.cell(row=i, column=2, value=value)
        c2.border = thin_border()
        c2.alignment = center_align()
        c2.fill = make_fill("EAF2F8" if i % 2 == 0 else "FFFFFF")
        
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    output_path = os.path.join(os.path.dirname(__file__), "Load_Testing_Report.xlsx")
    wb.save(output_path)
    print(f"Report saved to {output_path}")

if __name__ == "__main__":
    run_load_test()
